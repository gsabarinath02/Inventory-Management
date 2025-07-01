from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from datetime import date
from ...database import get_db
from ...core.crud import orders as orders_crud
from ...schemas.orders import OrderCreate, OrderUpdate, OrderResponse, OrderBulkCreate, OrderBulkResponse
from ...core.crud.audit_log import create_audit_log
from ...schemas.audit_log import AuditLogCreate
from ...api.deps import get_current_user
from ...schemas.user import User
from ...core.logging_context import current_user_var
import json
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import os
from pydantic import BaseModel
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from ...schemas.pending_order import PendingOrderCreate, PendingOrderUpdate
from ...core.crud import pending_order as pending_order_crud
from sqlalchemy import select
from app.models.pending_order import PendingOrder
from app.models.sales import SalesLog
import logging
from ...core.crud.orders import is_fully_delivered

router = APIRouter()
logger = logging.getLogger("orders-mirroring")

class OrderExportHeaders(BaseModel):
    party_name: str = ""
    destination: str = ""
    style: str = ""
    code: str = ""
    date: str = ""

@router.post("/orders/export-excel")
async def export_orders_excel(
    headers: OrderExportHeaders = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export all orders as an Excel file with logo and custom headers"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    orders = await orders_crud.get_all_orders(db, skip=0, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Insert logo (A1, merged A1:B5)
    logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../Backstitch-logo.png'))
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 120
        img.height = 120
        ws.add_image(img, 'A1')
    ws.merge_cells('A1:B5')

    # Merge and style header cells
    ws.merge_cells('C1:J1')
    ws['C1'] = 'Style'
    ws['C1'].font = Font(bold=True)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'] = headers.style
    ws['C3'] = headers.code
    ws.merge_cells('C2:J2')
    ws.merge_cells('C3:J3')
    ws['C2'].alignment = Alignment(horizontal='left')
    ws['C3'].alignment = Alignment(horizontal='left')

    ws.merge_cells('K1:L1')
    ws['K1'] = 'Party Name'
    ws['K1'].font = Font(bold=True)
    ws['K1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('K2:L2')
    ws['K2'] = headers.party_name
    ws['K2'].alignment = Alignment(horizontal='left')

    ws.merge_cells('K3:L3')
    ws['K3'] = 'Destination'
    ws['K3'].font = Font(bold=True)
    ws['K3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('K4:L4')
    ws['K4'] = headers.destination
    ws['K4'].alignment = Alignment(horizontal='left')

    ws.merge_cells('C4:J4')
    ws['C4'] = 'Transport - With Pass'
    ws['C4'].alignment = Alignment(horizontal='center')
    ws.merge_cells('K5:L5')
    ws['K5'] = 'Date'
    ws['K5'].font = Font(bold=True)
    ws['K5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('K6:L6')
    ws['K6'] = headers.date
    ws['K6'].alignment = Alignment(horizontal='left')

    # --- Table header ---
    size_columns = ['S', 'M', 'L', 'XL', 'XXL', '3XL', '4X', '5X']
    table_header = ["Color col", "Color"] + size_columns + ["Total"]
    ws.append(table_header)
    header_row = ws[7]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Table data ---
    row_num = 8
    for idx, order in enumerate(orders):
        sizes = order.sizes or {}
        row = [
            order.colour_code or '',
            order.color or '',
        ] + [sizes.get(size, 0) for size in size_columns] + [sum(sizes.get(size, 0) for size in size_columns)]
        ws.append(row)
        # Zebra striping
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') if idx % 2 == 0 else PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1

    # Set column widths
    for col, width in zip('ABCDEFGHIJKL', [10, 18, 6, 6, 6, 6, 6, 6, 6, 6, 10, 10]):
        ws.column_dimensions[col].width = width

    # Enable autofilter
    ws.auto_filter.ref = f"A7:L{row_num-1}"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="EXPORT_EXCEL",
            entity="Order",
            entity_id=0,
            field_changed="orders_export",
            new_value=f"Exported {len(orders)} orders to Excel"
        )
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orders-log.xlsx"}
    )

@router.post("/products/{product_id}/orders", response_model=OrderResponse)
async def create_order(
    product_id: int,
    order: OrderCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new order for a product"""
    if order.product_id != product_id:
        raise HTTPException(status_code=400, detail="Product ID mismatch")
    
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    db_order = await orders_crud.create_order(db, order)
    
    # Mirror to pending_orders
    pending_order = PendingOrderCreate(**order.model_dump())
    await pending_order_crud.create_pending_order(db, pending_order, db_order.order_number, db_order.financial_year)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="CREATE",
            entity="Order",
            entity_id=db_order.id,
            field_changed="order_number",
            new_value=str(db_order.order_number)
        )
    )
    
    return db_order

@router.get("/products/{product_id}/orders", response_model=List[OrderResponse])
async def get_orders(
    product_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    start_date: Optional[date] = Query(None, description="Start date for filtering (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="End date for filtering (YYYY-MM-DD)"),
    agency_name: Optional[str] = Query(None, description="Filter by agency name"),
    store_name: Optional[str] = Query(None, description="Filter by store name"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all orders for a product with optional filtering"""
    orders = await orders_crud.get_orders(
        db, 
        product_id, 
        skip=skip, 
        limit=limit,
        start_date=start_date,
        end_date=end_date,
        agency_name=agency_name,
        store_name=store_name
    )
    # Add fully_delivered to each order
    result = []
    for order in orders:
        order_dict = order.__dict__.copy()
        order_dict['fully_delivered'] = await is_fully_delivered(db, order)
        result.append(OrderResponse(**order_dict))
    return result

@router.get("/orders/", response_model=List[OrderResponse])
async def get_all_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all orders"""
    orders = await orders_crud.get_all_orders(db, skip=skip, limit=limit)
    return orders

@router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific order by ID"""
    order = await orders_crud.get_order(db, order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")
    return order

@router.put("/orders/{order_id}", response_model=OrderResponse)
async def update_order(
    order_id: int,
    order: OrderUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update an existing order with intelligent delivery logic"""
    # Set user context for audit logging
    current_user_var.set(current_user)

    # Fetch the original order
    orig_order = await orders_crud.get_order(db, order_id)
    if orig_order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    # Sum delivered quantities from SalesLog for this order_number and product_id
    delivered_logs = await db.execute(select(SalesLog).filter(SalesLog.order_number == orig_order.order_number, SalesLog.product_id == orig_order.product_id))
    delivered_logs = delivered_logs.scalars().all()
    delivered_total = {}
    for log in delivered_logs:
        for size, qty in (log.sizes or {}).items():
            delivered_total[size] = delivered_total.get(size, 0) + qty
    # Calculate updated order total per size
    updated_sizes = order.sizes or {}
    # Check if fully delivered
    fully_delivered = all((delivered_total.get(size, 0) >= orig_order.sizes.get(size, 0) for size in orig_order.sizes or {}))
    if fully_delivered:
        raise HTTPException(status_code=400, detail="This order has already been fully delivered. Please create a new order instead.")
    # Check if updated order total < delivered
    for size, delivered_qty in delivered_total.items():
        if updated_sizes.get(size, 0) < delivered_qty:
            raise HTTPException(status_code=400, detail=f"Cannot reduce {size} below delivered quantity ({delivered_qty}).")
    # Update the order
    updated_order = await orders_crud.update_order(db, order_id, order)
    if updated_order is None:
        raise HTTPException(status_code=404, detail="Order not found after update")
    # Mirror update to pending_orders (adjust pending quantities)
    result = await db.execute(select(PendingOrder).filter(PendingOrder.order_number == updated_order.order_number, PendingOrder.product_id == updated_order.product_id))
    db_pending_order = result.scalar_one_or_none()
    if db_pending_order:
        # Enforce: order.sizes[size] == pending_order.sizes[size] + delivered_total[size]
        new_pending = {}
        for size, updated_qty in updated_sizes.items():
            delivered_qty = delivered_total.get(size, 0)
            pending_qty = updated_qty - delivered_qty
            if pending_qty > 0:
                new_pending[size] = pending_qty
        logger.info(f"[INVARIANT-DEBUG] Order #{updated_order.order_number} sizes: {updated_sizes}")
        logger.info(f"[INVARIANT-DEBUG] Delivered totals: {delivered_total}")
        logger.info(f"[INVARIANT-DEBUG] New pending: {new_pending}")
        await pending_order_crud.update_pending_order(
            db,
            db_pending_order.id,
            PendingOrderUpdate(
                product_id=db_pending_order.product_id,
                color=db_pending_order.color,
                colour_code=db_pending_order.colour_code,
                sizes=new_pending,
                date=db_pending_order.date,
                agency_name=db_pending_order.agency_name,
                store_name=db_pending_order.store_name,
                operation=db_pending_order.operation or "Order"
            )
        )
        logger.info(f"Mirrored update to PendingOrder id={db_pending_order.id} for order_number={updated_order.order_number}")
    else:
        logger.warning(f"No PendingOrder found to mirror update for order_number={updated_order.order_number}")
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="UPDATE",
            entity="Order",
            entity_id=order_id,
            field_changed="order",
            new_value=str(order_id)
        )
    )
    # After updating, add fully_delivered to response
    updated_order_dict = updated_order.__dict__.copy()
    updated_order_dict['fully_delivered'] = await is_fully_delivered(db, updated_order)
    return OrderResponse(**updated_order_dict)

@router.delete("/orders/{order_id}")
async def delete_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an order"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    # Find order to get order_number and financial_year
    order = await orders_crud.get_order(db, order_id)
    success = await orders_crud.delete_order(db, order_id)
    if not success:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Mirror delete to pending_orders
    if order:
        result = await db.execute(select(PendingOrder).filter(PendingOrder.order_number == order.order_number, PendingOrder.financial_year == order.financial_year))
        db_pending_order = result.scalar_one_or_none()
        if db_pending_order:
            await pending_order_crud.delete_pending_order(db, db_pending_order.id)
            logger.info(f"Mirrored delete to PendingOrder id={db_pending_order.id} for order_number={order.order_number}, financial_year={order.financial_year}")
        else:
            logger.warning(f"No PendingOrder found to mirror delete for order_number={order.order_number}, financial_year={order.financial_year}")
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="DELETE",
            entity="Order",
            entity_id=order_id,
            field_changed="order",
            old_value=str(order_id)
        )
    )
    
    return {"message": "Order deleted successfully"}

@router.post("/products/{product_id}/orders/bulk", response_model=OrderBulkResponse)
async def create_orders_bulk(
    product_id: int,
    bulk_data: OrderBulkCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create multiple orders in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    result_orders = await orders_crud.create_orders_bulk(db, bulk_data.orders)
    # Mirror to pending_orders for each created order, using the actual created order fields
    for db_order in result_orders:
        pending_order = PendingOrderCreate(
            product_id=db_order.product_id,
            color=db_order.color,
            colour_code=db_order.colour_code,
            sizes=db_order.sizes,
            date=db_order.date,
            agency_name=db_order.agency_name,
            store_name=db_order.store_name,
            operation=db_order.operation or "Order"
        )
        await pending_order_crud.create_pending_order(db, pending_order, db_order.order_number, db_order.financial_year)
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_CREATE",
            entity="Order",
            entity_id=product_id,
            field_changed="bulk_orders",
            new_value=f"Created {len(result_orders)} orders"
        )
    )
    
    return OrderBulkResponse(rows_processed=len(result_orders), errors=None)

@router.delete("/products/{product_id}/orders/bulk")
async def delete_orders_bulk(
    product_id: int,
    date: date,
    agency_name: Optional[str] = None,
    store_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete multiple orders in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    deleted_count = await orders_crud.delete_orders_bulk(db, date, agency_name, store_name)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_DELETE",
            entity="Order",
            entity_id=product_id,
            field_changed="bulk_orders",
            old_value=f"Deleted {deleted_count} orders for date {date}"
        )
    )
    
    return {"message": f"Deleted {deleted_count} orders", "deleted_count": deleted_count} 