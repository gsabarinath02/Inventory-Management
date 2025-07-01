from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from datetime import date
from ...database import get_db
from ...core.crud import sales as sales_crud
from ...schemas.sales import SalesLogCreate, SalesLogUpdate, SalesLog
from ...core.crud.audit_log import create_audit_log
from ...schemas.audit_log import AuditLogCreate
from ...api.deps import get_current_user
from ...schemas.user import User
from ...core.logging_context import current_user_var
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import io, os
from ...core.crud.audit_log import create_audit_log
from ...schemas.audit_log import AuditLogCreate
from pydantic import BaseModel

router = APIRouter()

# Legacy routes for frontend compatibility
@router.get("/{product_id}", response_model=List[SalesLog])
async def get_sales_logs_legacy(
    product_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    agency_name: Optional[str] = Query(None),
    store_name: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Get sales logs for a product"""
    # If single date and store_name are provided, get the last matching entry
    if date and store_name:
        return await sales_crud.get_last_sales_log_by_date_and_store(db, product_id=product_id, date=date, store_name=store_name)
    
    return await sales_crud.get_sales_logs_by_product(db, product_id=product_id, start_date=start_date, end_date=end_date, agency_name=agency_name, store_name=store_name)

@router.post("/", response_model=SalesLog)
async def create_sales_log_legacy(
    sales_log: SalesLogCreate, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Create a sales log"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    db_log = await sales_crud.create_sales_log(db=db, sales_log=sales_log)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="CREATE",
            entity="SalesLog",
            entity_id=db_log.id,
            field_changed="sales_log",
            new_value=str(db_log.id)
        )
    )
    
    return db_log

@router.post("/bulk-create", response_model=List[SalesLog])
async def create_sales_logs_bulk_legacy(
    sales_logs: List[SalesLogCreate], 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Create multiple sales logs in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    result = await sales_crud.create_sales_logs_bulk(db=db, sales_logs=sales_logs)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_CREATE",
            entity="SalesLog",
            entity_id=0,
            field_changed="bulk_sales_logs",
            new_value=f"Created {len(sales_logs)} sales logs"
        )
    )
    
    return result

@router.delete("/bulk-delete")
async def delete_sales_logs_bulk_legacy(
    product_id: int,
    date: Optional[str] = Query(None),
    store_name: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Delete multiple sales logs in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    deleted_count = await sales_crud.delete_sales_logs_bulk(db=db, product_id=product_id, date=date, store_name=store_name)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_DELETE",
            entity="SalesLog",
            entity_id=0,
            field_changed="bulk_sales_logs",
            old_value=f"Deleted {deleted_count} sales logs for product {product_id}"
        )
    )
    
    return {"message": f"Deleted {deleted_count} sales log entries"}

@router.put("/{log_id}", response_model=SalesLog)
async def update_sales_log_legacy(
    log_id: int, 
    sales_log: SalesLogUpdate, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Update a sales log"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    # Get old log for audit
    old_log = await sales_crud.get_sales_log_by_id(db, log_id)
    db_log = await sales_crud.update_sales_log(db=db, log_id=log_id, sales_log=sales_log)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="UPDATE",
            entity="SalesLog",
            entity_id=log_id,
            field_changed="sales_log",
            old_value=str(old_log.id) if old_log else None,
            new_value=str(db_log.id) if db_log else None
        )
    )
    
    if db_log is None:
        raise HTTPException(status_code=404, detail="Sales log not found")
    return db_log

@router.delete("/{log_id}", response_model=SalesLog)
async def delete_sales_log_legacy(
    log_id: int, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Legacy route for frontend compatibility - Delete a sales log"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    db_log = await sales_crud.delete_sales_log(db=db, log_id=log_id)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="DELETE",
            entity="SalesLog",
            entity_id=log_id,
            field_changed="sales_log",
            old_value=str(db_log.id) if db_log else None,
            new_value=None
        )
    )
    
    if db_log is None:
        raise HTTPException(status_code=404, detail="Sales log not found")
    return db_log

# New structured routes
@router.post("/{product_id}/sales", response_model=SalesLog)
async def create_sales_log(
    product_id: int,
    sales_log: SalesLogCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new sales log for a product"""
    if sales_log.product_id != product_id:
        raise HTTPException(status_code=400, detail="Product ID mismatch")
    
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    db_sales_log = await sales_crud.create_sales_log(db, sales_log)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="CREATE",
            entity="SalesLog",
            entity_id=db_sales_log.id,
            field_changed="sales_log",
            new_value=str(db_sales_log.id)
        )
    )
    
    return db_sales_log

@router.get("/{product_id}/sales", response_model=List[SalesLog])
async def get_sales_logs(
    product_id: int,
    start_date: Optional[str] = Query(None, description="Start date for filtering (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for filtering (YYYY-MM-DD)"),
    store_name: Optional[str] = Query(None, description="Filter by store name"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all sales logs for a product with optional filtering"""
    sales_logs = await sales_crud.get_sales_logs_by_product(
        db, 
        product_id, 
        start_date=start_date,
        end_date=end_date,
        store_name=store_name
    )
    return sales_logs

@router.put("/{product_id}/sales/{sales_log_id}", response_model=SalesLog)
async def update_sales_log(
    product_id: int,
    sales_log_id: int,
    sales_log: SalesLogUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update an existing sales log"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    updated_sales_log = await sales_crud.update_sales_log(db, sales_log_id, sales_log)
    if updated_sales_log is None:
        raise HTTPException(status_code=404, detail="Sales log not found")
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="UPDATE",
            entity="SalesLog",
            entity_id=sales_log_id,
            field_changed="sales_log",
            new_value=str(sales_log_id)
        )
    )
    
    return updated_sales_log

@router.delete("/{product_id}/sales/{sales_log_id}")
async def delete_sales_log(
    product_id: int,
    sales_log_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a sales log"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    deleted_log = await sales_crud.delete_sales_log(db, sales_log_id)
    if deleted_log is None:
        raise HTTPException(status_code=404, detail="Sales log not found")
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="DELETE",
            entity="SalesLog",
            entity_id=sales_log_id,
            field_changed="sales_log",
            old_value=str(sales_log_id)
        )
    )
    
    return {"message": "Sales log deleted successfully"}

@router.post("/{product_id}/sales/bulk", response_model=List[SalesLog])
async def create_sales_logs_bulk(
    product_id: int,
    sales_logs: List[SalesLogCreate],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create multiple sales logs in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    result = await sales_crud.create_sales_logs_bulk(db, sales_logs)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_CREATE",
            entity="SalesLog",
            entity_id=product_id,
            field_changed="bulk_sales_logs",
            new_value=f"Created {len(result)} sales logs"
        )
    )
    
    return result

@router.delete("/{product_id}/sales/bulk")
async def delete_sales_logs_bulk(
    product_id: int,
    date: Optional[str] = Query(None, description="Date for filtering (YYYY-MM-DD)"),
    store_name: Optional[str] = Query(None, description="Filter by store name"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete multiple sales logs in bulk"""
    # Set user context for audit logging
    current_user_var.set(current_user)
    
    deleted_count = await sales_crud.delete_sales_logs_bulk(db, product_id, date, store_name)
    
    # Log the audit event
    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="BULK_DELETE",
            entity="SalesLog",
            entity_id=product_id,
            field_changed="bulk_sales_logs",
            old_value=f"Deleted {deleted_count} sales logs for date {date}"
        )
    )
    
    return {"message": f"Deleted {deleted_count} sales logs", "deleted_count": deleted_count}

class SalesExportHeaders(BaseModel):
    party_name: str = ""
    destination: str = ""
    style: str = ""
    code: str = ""
    date: str = ""

@router.post("/export-excel")
async def export_sales_excel(
    headers: SalesExportHeaders = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export all sales logs as an Excel file with logo and custom headers"""
    current_user_var.set(current_user)
    sales_logs = await sales_crud.get_all_sales_logs(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Logo
    logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../Backstitch-logo.png'))
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 120
        img.height = 120
        ws.add_image(img, 'A1')
    ws.merge_cells('A1:B5')

    # Header rows (merged cells)
    ws.merge_cells('C1:J1')
    ws.merge_cells('K1:N1')
    ws.merge_cells('O1:R1')
    ws['C1'] = 'Style'
    ws['C1'].font = Font(bold=True, size=20)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K1'] = 'Party Name'
    ws['K1'].font = Font(bold=True, size=20)
    ws['K1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O1'] = 'Destination'
    ws['O1'].font = Font(bold=True, size=20)
    ws['O1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C2:J2')
    ws.merge_cells('K2:N2')
    ws.merge_cells('O2:R2')
    ws['C2'] = ''
    ws['K2'] = ''
    ws['O2'] = ''

    ws.merge_cells('C3:R3')
    ws['C3'] = 'Transport - With Pass'
    ws['C3'].font = Font(bold=True, size=16)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C4:N4')
    ws.merge_cells('O4:P4')
    ws.merge_cells('Q4:R4')
    ws['O4'] = 'Date'
    ws['O4'].font = Font(bold=True)
    ws['O4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Q4'] = headers.date
    ws['Q4'].alignment = Alignment(horizontal='center', vertical='center')

    # Table header
    table_header = ["Color col", "Color", "s", "m", "l", "Total"]
    ws.append([None, None] + table_header + [None]*10)
    header_row = ws[6]
    for cell in header_row[2:8]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Data rows
    row_num = 7
    for idx, log in enumerate(sales_logs):
        sizes = log.sizes or {}
        row = [
            log.colour_code or '',
            log.color or '',
            sizes.get('s', 0),
            sizes.get('m', 0),
            sizes.get('l', 0),
            (sizes.get('s', 0) + sizes.get('m', 0) + sizes.get('l', 0)),
        ]
        ws.append([None, None] + row + [None]*10)
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') if idx % 2 == 0 else PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for col in range(3, 9):
            cell = ws.cell(row=row_num+1, column=col)
            cell.fill = fill
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    # Set column widths
    for col, width in zip('ABCDEFGHIJKLMN', [6, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    await create_audit_log(
        db,
        AuditLogCreate(
            user_id=current_user.id,
            username=current_user.email,
            action="EXPORT_EXCEL",
            entity="SalesLog",
            entity_id=0,
            field_changed="sales_export",
            new_value=f"Exported {len(sales_logs)} sales logs to Excel"
        )
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales-log.xlsx"}
    ) 