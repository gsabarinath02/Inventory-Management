from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from datetime import date
from ...database import get_db
from ...core.crud import pending_order as pending_order_crud
from ...schemas.pending_order import PendingOrderCreate, PendingOrderUpdate, PendingOrderResponse
from ...api.deps import get_current_user
from ...schemas.user import User
from ...core.logging_context import current_user_var
from ...schemas.sales import SalesLogCreate
from ...core.crud import sales as sales_crud
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import io, os
from pydantic import BaseModel

router = APIRouter()

@router.get("/products/{product_id}/pending-orders", response_model=List[PendingOrderResponse])
async def get_pending_orders(
    product_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return await pending_order_crud.get_pending_orders(db, product_id, skip=skip, limit=limit)

@router.post("/pending-orders/{pending_order_id}/deliver")
async def deliver_pending_order(
    pending_order_id: int,
    delivered_sizes: dict = Body(...),
    delivery_date: str = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Deliver a pending order (fully or partially). If fully delivered, move to sales log and remove from pending orders. If partially, move delivered part to sales log and update pending order.
    """
    pending_order = await pending_order_crud.get_pending_order_by_id(db, pending_order_id)
    if not pending_order:
        raise HTTPException(status_code=404, detail="Pending order not found")
    # Call new deliver_pending_order logic in CRUD
    result = await pending_order_crud.deliver_pending_order(db, pending_order, delivered_sizes, delivery_date)
    return result

class PendingOrdersExportHeaders(BaseModel):
    party_name: str = ""
    destination: str = ""
    style: str = ""
    code: str = ""
    date: str = ""

@router.post("/export-excel")
async def export_pending_orders_excel(
    headers: PendingOrdersExportHeaders = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    current_user_var.set(current_user)
    pending_orders = await pending_order_crud.get_all_pending_orders(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Orders"

    # Logo
    logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../Backstitch-logo.png'))
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 120
        img.height = 120
        ws.add_image(img, 'A1')
    ws.merge_cells('A1:B5')

    # Header rows (merged cells)
    ws.merge_cells('C1:J1')
    ws.merge_cells('K1:N1')
    ws.merge_cells('O1:R1')
    ws['C1'] = 'Style'
    ws['C1'].font = Font(bold=True, size=20)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K1'] = 'Party Name'
    ws['K1'].font = Font(bold=True, size=20)
    ws['K1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O1'] = 'Destination'
    ws['O1'].font = Font(bold=True, size=20)
    ws['O1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C2:J2')
    ws.merge_cells('K2:N2')
    ws.merge_cells('O2:R2')
    ws['C2'] = ''
    ws['K2'] = ''
    ws['O2'] = ''

    ws.merge_cells('C3:R3')
    ws['C3'] = 'Transport - With Pass'
    ws['C3'].font = Font(bold=True, size=16)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C4:N4')
    ws.merge_cells('O4:P4')
    ws.merge_cells('Q4:R4')
    ws['O4'] = 'Date'
    ws['O4'].font = Font(bold=True)
    ws['O4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Q4'] = headers.date
    ws['Q4'].alignment = Alignment(horizontal='center', vertical='center')

    # Table header
    table_header = ["Color col", "Color", "s", "m", "l", "Total"]
    ws.append([None, None] + table_header + [None]*10)
    header_row = ws[6]
    for cell in header_row[2:8]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Data rows
    row_num = 7
    for idx, log in enumerate(pending_orders):
        sizes = log.sizes or {}
        row = [
            log.colour_code or '',
            log.color or '',
            sizes.get('s', 0),
            sizes.get('m', 0),
            sizes.get('l', 0),
            (sizes.get('s', 0) + sizes.get('m', 0) + sizes.get('l', 0)),
        ]
        ws.append([None, None] + row + [None]*10)
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') if idx % 2 == 0 else PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        for col in range(3, 9):
            cell = ws.cell(row=row_num+1, column=col)
            cell.fill = fill
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    # Set column widths
    for col, width in zip('ABCDEFGHIJKLMN', [6, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pending-orders.xlsx"}
    )

# Additional endpoints for create, update, delete can be added as needed 